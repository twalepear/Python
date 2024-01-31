print('Thoa Le')
print('o----')
print(' ||||')  # supposed to be a dog and show us that there is only one character per space
print('*' * 10)  # is an expression of 10 asterisks
# Should I spend 2 hours everyday learning Python every for 3 months?
print('VARIABLES')
price = 10  # price is a label for a variable, in this case 10. This an integer
price = 20  # the last time that the line rewrites the variable, will execute that command
rating = 4.9  # numbers with decimal points are floats
name = 'Thoa'  # this is a string
is_published = True  # use underscore as spaces in variables. Use uppercase for T/F. This is a Boolean
print(price)

name = 'John Smith'  # don't forget about the apostrophes
age = 20
new_patient = True
print('Name:' + name)
print('Age:' + str(age))  # when there is a string and an integer in the same line, need to convert one to another
print('New patient:' + str(new_patient))  # the same with string and Boolean terms

name = input('What is your name? ')  # this gives us a string
colour = input('What is your favourite colour? ')
print(name + ' likes ' + colour)

birth_year = input()  # string and integers

first = 'John'
last = 'Smith'
message = first + '[' + last + '] is a coder'  # should print(message) -> John [Smith] is a coder
msg = f'{first} [{last}] is a coder'  # should also print(msg) -> John [Smith] is a coder

course = 'Python for Beginners'
print(course[-2])  # index [0:3] will print Pyt
test = 'Jennifer'
print(test[1:-1])  # will print ennife

print(len(course))  # gives length (number of characters) of the string
print(course.upper())  # this is a method of a string -> makes the string uppercase; same with lower to lowercase

import math  # allows you to use maths functions

print(math.floor(2.9))  # one of the many maths functions

is_hot = False
is_cold = True  # change these to activate the print statements

if is_hot:
    print("It's a hot day")
    print("Drink plenty of water")
elif is_cold:
    print("It's a cold day")
    print("Wear warm clothes")
else:
    print("It's a lovely day")
print("Enjoy your day")

price = 1000000
has_good_credit = False

if has_good_credit:
    down_payment = 0.1 * price
else:
    down_payment = 0.1 * price
print(f"Down payment: ${int(down_payment)}")

has_high_income = True
has_good_credit = True

if has_high_income and has_good_credit:  # can use OR if the condition only needs to be one of them to be true
    print("Eligible for loan")

# can use BMI as exercise for comparison operators < and > , make sure to use == for comparison to avoid confusion

name = input("What is your name? ")

if len(name) < 3:
    print('name must be at least 3 characters')
elif len(name) > 50:
    print('name can be a maximum of 50 characters')
else:
    print('name looks good')

print("Type 'help' to get menu options")
input_command = ""
car_started = False

while True:
    input_command = input(">").lower()
    if input_command == "help":
        print("""
start - to start the car
stop - to stop the car
quit - to exit""")
    elif input_command == "start":
        if not car_started:
            print("Car started... Ready to go!")
            car_started = True
        elif car_started:
            print("Car is already on!")
    elif input_command == "stop":
        if not car_started:
            print("what are we stopping here? To quit the game type 'quit'.")
        elif car_started:
            print("Car stopped.")
            car_started = False
    elif input_command == "quit":
        print("You have quit the game. Goodbye.")
        break
    else:
        print("I don't understand that...")

prices = [10, 20, 30]

total = 0
for price in prices:
    total += price
print(f"Total: {total}")

for x in range(4):
    for y in range(3):
        print(f'{x},{y})')

numbers = [5, 2, 5, 2, 2]

for items in numbers:
    text = ""
    for number_of_x in range(items):
        text += "x"
    print(text)

list = [4, 2, 5, 7, 9, 3]
found_highest_number = list[0]

for numbers in list:
    if numbers > found_highest_number:
        found_highest_number = numbers
print(found_highest_number)

matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
print(matrix[0][1])  # should get 2

number_list = [2, 3, 2, 4, 5, 8, 2, 6, 4]

for item in number_list:
    counter = 0
    for compared_number in number_list:
        if item == compared_number:
            counter += 1
            if counter > 1:
                number_list.remove(item)
                counter -= 1

print(number_list)

number_list = [2, 3, 2, 4, 5, 8, 2, 6, 4]
uniques =[]
for item in number_list:
    if item not in uniques:
        uniques.append(item)
print(uniques)

numbers = {
    "1": "one",
    "2": "two",
    "3": "three",
    "4": "four",
    "5": "five",
    "6": "six",
    "7": "seven",
    "8": "eight",
    "9": "nine",
    "0": "zero"
}

phone = input("Phone: ")

output = ""
for number in phone:
    output += numbers.get(number, "!") + " "
print(output)

message = input(">")
words = message.split(' ')
emojis = {
    ":)": "😀",
    ":(": "🙁"
}
output = ""
for word in words:
    output += emojis.get(word, word) + " "
print(output)


def greet_user():
    print('Hi there!')
    print('Welcome aboard')


print("Start")
greet_user()
print("Finish")

def greet_user(first_name, last_name):
    print(f'Hi {first_name} {last_name}!')
    print('Welcome aboard')


print("Start")
greet_user("John", "Smith")  # need to supply a name otherwise there will be an error
greet_user(last_name="Smith", first_name="John")  # keyword arguments improve readability esp numbers
# if using mixture, start with "John" (first position) then any keyword argument is fine
print("Finish")

def square(number):
    return number * number  # need the return function so it doesn't return none

square(3)  # doesn't do anything
print(square(3))  # will print 9

def output_emoji(message):
    words = message.split(' ')
    emojis = {
        ":)": "😀",
        ":(": "🙁"
    }
    output = ""
    for word in words:
        output += emojis.get(word, word) + " "
    return output


message = input(">")
print(output_emoji(message))

try:
    age = int(input('Age: '))
    income = 20000
    risk = income / age
    print(age)
except ZeroDivisionError:
    print('Age cannot be 0.')
except ValueError:
    print('Invalid value')

class Point:  # uppercase based on Pascal naming convention
    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point()
point1.x = 10
point1.y = 20
print(point1.x)
point1.draw()

point2 = Point()
point2.x = 20
print(point2.x)


class Point:
    def __init__(self, x, y):  # constructors
        self.x = x
        self.y = y

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point = Point(10, 20)
point.x = 11
print(point.x)

class Person:
    def __init__(self, name):
        self.name = name

    def talk(self):
        print(f"Hi, I am {self.name}")


john = Person("John Smith")
john.talk()

bob = Person("Bob Smith")
bob.talk()


class Dog:
    def walk(self):
        print("walk")


class Cat:
    def walk(self):
        print("walk")


# DRY: DON'T REPEAT YOURSELF

class Mammal:
    def walk(self):
        print("walk")


class Dog(Mammal):
    def bark(self):
        print("bark")


class Cat(Mammal)
    pass


dog1 = Dog()
dog1.walk()

def lbs_to_kg(weight):  # copy this and paste into new py (name it converters,etc)
    return weight * 0.45


def kg_to_lbs(weight):
    return weight / 0.45

import converters
from converters import kg_to_lbs

kg_to_lbs(100)

print(converters.kg_to_lbs(70))

import utils

print(utils.find_max([2,4,8,4,3,5,7]))

import ecommerce.shipping
ecommerce.shipping.calc_shipping()

import random

for i in range(3):
    print(random.randint(10, 20))

members = ['John', 'Mary', 'Bob', 'Mosh']
leader = random.choice(members)
print(leader)

import random


class Dice:
    def roll(self):
        first = random.randint(1,6)
        second = random.randint(1, 6)
        return first, second


dice = Dice()
print(dice.roll())

from pathlib import Path  # Absolute path c:\Program Files\ or Relative path

path = Path("ecommerce")
print(path.exists())

path = Path("emails")
print(path.mkdir())  # make directory
print(path.rmdir())  # remove directory

print(path.glob('*.*'))  # search function and parameter means to include everything

path = Path()
for file in path.glob('*.py'):  # to show all the py files
    print(file)

import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']  # cell = sheet['a1'] is the same thing as "cell = sheet.cell(1,1)"

    for row in range(2, sheet.max_row + 1):  # (1,4) will return 1,2,3 but not 4
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)

# 1. Import the Data
# 2. Clean the Data
# 3. Split the Data into Training/Data Sets
# 4. Create a Model
# 5. Train the Model
# 6. Make Predictions
# 7. Evaluate and Improve
